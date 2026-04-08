#!/usr/bin/env python3
import argparse
import csv
import io
import json
import re
from collections import Counter
from datetime import datetime, timezone
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = Path("/Users/victoria.serrano/Downloads/Levels - feed the bear (1).xlsx")
DEFAULT_OUTPUT = ROOT / "output" / "spreadsheet" / "Levels_feed_the_bear_after_feedback_sync.xlsx"
PLAYTEST_EVENTS_PATH = ROOT / "playtest" / "playtest_events.jsonl"
PLAYTEST_LATEST_SESSION_PATH = ROOT / "playtest" / "latest_play_session.json"
PLAYTEST_SUMMARY_PATH = ROOT / "output" / "playtest" / "playtest_summary.json"
README_SHEET_NAME = "README"
PRIMARY_SHEET_NAME = "All Progressions"
PRIMARY_HEADERS = [
    "Position in Progression",
    "Level ID",
    "Name/Title",
    "Progression",
    "Difficulty",
    "Screenshot",
    "Code File",
    "Need tweak",
    "Need fixes",
    "Feedback",
    "Feedback Owner",
    "Playtest difficulty",
    "Notes",
]
SCREENSHOT_HEADERS = [
    "Level Name",
    "Screenshot",
    "Code File",
    "Local Screenshot Path",
    "Status",
]
DB_HEADERS = [
    "row_type",
    "saved_at",
    "item_id",
    "progression_key",
    "progression_label",
    "slot",
    "level_file",
    "level_name",
    "difficulty",
    "board",
    "pairs",
    "blockers",
    "moves",
    "status",
    "changed",
    "notes",
    "source_path",
    "saved_path",
]
CATALOG_SHEET_NAME = "Level Catalog"
CATALOG_HEADERS = [
    "File",
    "Display Name",
    "Source Progression",
    "Difficulty",
    "Difficulty Score",
    "Board",
    "Pairs",
    "Blockers",
    "Moves",
    "Status",
    "Screenshot",
    "Saved Path",
]
PLANNER_SHEET_NAME = "Mix Planner"
PLANNER_SLOT_HEADERS = [f"Slot {index}" for index in range(1, 11)]
PLANNER_HEADERS = [
    "Proposal ID",
    "Folder Name",
    "Approved",
    "Materialized",
    "Source Progression",
    "Suggestion",
    "Tutorial File",
    *PLANNER_SLOT_HEADERS,
    "Curve",
    "Notes",
    "Output Folder",
]
PLANNER_ROW_BLOCK_SIZE = 3
RENAME_SHEET_NAME = "Level Renames"
RENAME_HEADERS = [
    "Occurrence ID",
    "Progression",
    "Position",
    "Ordinal",
    "Screenshot",
    "Current Name",
    "Target Name",
    "Preview Label",
    "Code File",
    "Level Path",
    "Screenshot File",
    "Screenshot Path",
    "Planned File",
    "Rename Pending",
    "Apply Status",
    "Notes",
]
README_HEADERS = ["Group", "Tab / Surface", "What it does", "Edit Rule", "Link"]
TOOLKIT_BASE_URL = "http://127.0.0.1:8080"
HEADER_FILL = PatternFill("solid", fgColor="DCEAF7")
DIFFICULTY_FILLS = {
    "EASY": PatternFill("solid", fgColor="DFF3E4"),
    "MEDIUM": PatternFill("solid", fgColor="FFF1CC"),
    "HARD": PatternFill("solid", fgColor="F8D7DA"),
}
DIFFICULTY_FONT_COLORS = {
    "EASY": "245C3B",
    "MEDIUM": "8A5A00",
    "HARD": "8A1F2D",
}
DIFFICULTY_DEFAULT_FILL = PatternFill("solid", fgColor="E7EDF1")
DIFFICULTY_DEFAULT_FONT_COLOR = "4E5D6C"
RENAME_LOG_PATH = ROOT / "output" / "spreadsheet" / "level_rename_log.csv"
WORKBOOK_TAB_COLORS = {
    "read": "4E79A7",
    "interactive": "59A14F",
    "data": "F28E2B",
}


def load_json(path: Path):
    return json.loads(path.read_text())


def load_jsonl(path: Path):
    rows = []
    if not path.is_file():
        return rows
    for index, line in enumerate(path.read_text(encoding="utf8").splitlines(), start=1):
        text = line.strip()
        if not text:
            continue
        try:
            rows.append(json.loads(text))
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"Failed to parse JSONL line {index} in {path}: {exc}") from exc
    return rows


def to_number(value, fallback=0):
    try:
        num = float(value)
    except (TypeError, ValueError):
        return fallback
    return num if num == num else fallback


def safe_str(value):
    if value is None:
        return ""
    return str(value)


def resolve_project_path(path_value: str) -> Path:
    raw = safe_str(path_value).strip()
    if not raw:
        return Path()
    candidate = Path(raw)
    if candidate.is_absolute():
        return candidate
    if raw.startswith(f"{ROOT.name}/"):
        return ROOT.parent / raw
    return ROOT / raw


def display_level_name(file_name: str, screenshot_name: str = "") -> str:
    stem = Path(screenshot_name or file_name).stem
    stem = stem.replace("tutorial_level", "tutorial")
    stem = stem.replace("lvl_", "", 1) if stem.startswith("lvl_") else stem
    parts = stem.split("_", 1)
    if len(parts) == 2 and parts[0].isdigit():
        stem = parts[1]
    stem = stem.replace("_", " ").replace("-", " ")
    return " ".join(part for part in stem.split() if part)


def ordinal_label(slot_value) -> str:
    try:
        return f"{int(str(slot_value).strip() or '0'):02d}"
    except (TypeError, ValueError):
        return safe_str(slot_value).strip()


def prefixed_level_label(slot_value, name: str) -> str:
    ordinal = ordinal_label(slot_value)
    label = safe_str(name).strip()
    if ordinal and label:
        return f"{ordinal} · {label}"
    return label or ordinal


def load_screenshot_lookup() -> dict[str, dict]:
    lookup = {}
    if not RENAME_LOG_PATH.exists():
        return lookup
    with RENAME_LOG_PATH.open(newline="", encoding="utf8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            screenshot_file = safe_str(row.get("screenshot_file")).strip()
            if not screenshot_file:
                continue
            screenshot_path = ROOT / "levels" / "screenshots" / screenshot_file
            keys = {
                safe_str(row.get("display_code_file")).strip(),
                safe_str(row.get("canonical_json")).strip(),
                safe_str(row.get("source_file")).strip(),
            }
            aliases = safe_str(row.get("aliases")).split("|")
            for alias in aliases:
                keys.add(alias.strip())
            for key in keys:
                if not key:
                    continue
                if key in lookup:
                    continue
                lookup[key] = {
                    "screenshot_name": screenshot_file,
                    "screenshot_path": str(screenshot_path),
                }
    return lookup


def infer_screenshot_for_level(level_file: str, screenshot_lookup: dict[str, dict]) -> tuple[str, str]:
    alias = screenshot_lookup.get(level_file) or screenshot_lookup.get(Path(level_file).name)
    if alias:
        return alias["screenshot_name"], alias["screenshot_path"]
    stem = Path(level_file).stem
    candidates = [
        ROOT / "levels" / "screenshots" / f"{stem}.png",
        ROOT / "screenshots" / f"{stem}.png",
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate.name, str(candidate)
    return "", ""


def title_case_difficulty(value: str) -> str:
    text = safe_str(value).strip()
    return text[:1].upper() + text[1:].lower() if text else ""


def compact_progression_label(label: str) -> str:
    text = safe_str(label).strip()
    lower = text.lower()
    # Preserve explicit version labels like "Progression A v4".
    match = re.match(r"^progression\s+([a-f])(?:\s+(v\d+))?\s*$", lower)
    if match:
        letter = match.group(1).upper()
        version = match.group(2)
        return f"Progression {letter} {version}" if version else f"Progression {letter}"
    if lower.startswith("progression ") and len(text) > len("Progression "):
        # Legacy compacting for labels like "Progression 4" or "Progression D".
        return f"Progression {text.split()[-1].upper()}"
    return text


def progression_family_code(value: str) -> str:
    text = safe_str(value).strip()
    if not text:
        return ""
    lowered = text.lower()
    match = re.match(r"^progression\s+([a-f])(?:\s+v\d+)?\s*$", lowered)
    if match:
        return match.group(1).upper()
    if lowered.endswith("a") or lowered.endswith(" a"):
        return "A"
    if lowered.endswith("b") or lowered.endswith(" b"):
        return "B"
    if lowered.endswith("c") or lowered.endswith(" c"):
        return "C"
    if "progressiona" in lowered or "progression a" in lowered:
        return "A"
    if "progressionb" in lowered or "progression b" in lowered:
        return "B"
    if "progressionc" in lowered or "progression c" in lowered:
        return "C"
    return ""


def difficulty_score(value: str, moves=0, pairs=0, blockers=0) -> float:
    base = {
        "EASY": 1.0,
        "MEDIUM": 2.0,
        "HARD": 3.0,
    }.get(safe_str(value).strip().upper(), 0.0)
    try:
        moves_value = float(moves or 0)
    except (TypeError, ValueError):
        moves_value = 0.0
    try:
        pairs_value = float(pairs or 0)
    except (TypeError, ValueError):
        pairs_value = 0.0
    try:
        blockers_value = float(blockers or 0)
    except (TypeError, ValueError):
        blockers_value = 0.0
    return round(base + (moves_value * 0.01) + (pairs_value * 0.001) + (blockers_value * 0.0001), 3)


def item_index(snapshot: dict) -> dict:
    return {item.get("id"): item for item in snapshot.get("items", [])}


def find_slot_item(slot_entry: dict, items_by_id: dict) -> dict:
    item_id = slot_entry.get("item_id")
    return items_by_id.get(item_id) or {}


def build_primary_rows(snapshot: dict):
    rows = []
    screenshot_rows = []
    sequence = 1
    items_by_id = item_index(snapshot)
    screenshot_lookup = load_screenshot_lookup()
    for progression_key in snapshot.get("progression_order", []):
        progression = (snapshot.get("progressions") or {}).get(progression_key) or {}
        progression_label = compact_progression_label(progression.get("label", progression_key))
        for slot_entry in progression.get("slots", []):
            level_file = safe_str(slot_entry.get("file")).strip()
            if not level_file:
                continue
            item = find_slot_item(slot_entry, items_by_id)
            screenshot_path = safe_str(item.get("screenshot_path")).strip()
            resolved_item_path = resolve_project_path(screenshot_path)
            if resolved_item_path.is_file():
                screenshot_path = str(resolved_item_path)
            else:
                screenshot_path = ""
            screenshot_name = Path(screenshot_path).name if screenshot_path else ""
            if not screenshot_name:
                screenshot_name, screenshot_path = infer_screenshot_for_level(level_file, screenshot_lookup)
            difficulty = title_case_difficulty(slot_entry.get("difficulty") or item.get("difficulty"))
            name = display_level_name(level_file, screenshot_name)
            level_id = f"{sequence:03d}"
            note = safe_str(slot_entry.get("notes"))
            primary = {
                "Position in Progression": slot_entry.get("slot", ""),
                "Level ID": level_id,
                "Name/Title": prefixed_level_label(slot_entry.get("slot", ""), name),
                "Progression": progression_label,
                "Difficulty": difficulty,
                "Screenshot": screenshot_name,
                "Code File": level_file,
                "Need tweak": "yes" if item.get("changed") else "no",
                "Need fixes": "FALSE",
                "Feedback": "",
                "Feedback Owner": "",
                "Playtest difficulty": "",
                "Notes": note,
                "_screenshot_path": screenshot_path,
            }
            rows.append(primary)
            screenshot_rows.append({
                "Level Name": name,
                "Screenshot": screenshot_name,
                "Code File": level_file,
                "Local Screenshot Path": screenshot_path,
                "Status": "FOUND" if screenshot_path and Path(screenshot_path).exists() else "MISSING",
                "_screenshot_path": screenshot_path,
            })
            sequence += 1
    return rows, screenshot_rows


def board_label(level: dict | None) -> str:
    level = level or {}
    width = level.get("board_width") or level.get("width") or ""
    height = level.get("board_height") or level.get("height") or ""
    if width and height:
        return f"{width}x{height}"
    return ""


def tweak_reasons_for_record(record: dict | None) -> list[str]:
    record = record or {}
    reasons = []
    validation = safe_str(record.get("validation_status")).upper()
    solved = bool(record.get("solved"))
    save_reason = safe_str(record.get("save_reason")).lower()
    history_length = int(to_number(record.get("history_length"), 0))
    if validation and validation != "OK":
        reasons.append("validation_invalid")
    if not solved:
        reasons.append("unsolved")
    if save_reason == "validate" and not solved:
        reasons.append("validate_unsolved")
    if history_length >= 40:
        reasons.append("long_history")
    return reasons


def normalize_playtest_record(record: dict | None) -> dict:
    record = record or {}
    session = record.get("session") if isinstance(record.get("session"), dict) else {}
    level = session.get("level") if isinstance(session, dict) and isinstance(session.get("level"), dict) else {}
    history = session.get("history") if isinstance(session, dict) and isinstance(session.get("history"), list) else []
    path_lengths = record.get("path_lengths") if isinstance(record.get("path_lengths"), dict) else {}
    level_pairs = level.get("pairs") if isinstance(level.get("pairs"), list) else []
    level_blockers = level.get("blockers") if isinstance(level.get("blockers"), list) else []
    return {
        "saved_at": safe_str(record.get("saved_at")),
        "origin": safe_str(record.get("origin")),
        "save_reason": safe_str(record.get("save_reason")),
        "solved": bool(record.get("solved")),
        "level_name": safe_str(record.get("level_name")),
        "level_number": record.get("level_number") if record.get("level_number") is not None else "",
        "board": safe_str(record.get("board")) or board_label(level),
        "board_width": record.get("board_width") if record.get("board_width") is not None else level.get("board_width", ""),
        "board_height": record.get("board_height") if record.get("board_height") is not None else level.get("board_height", ""),
        "pairs": record.get("pairs") if record.get("pairs") is not None else len(level_pairs),
        "blockers": record.get("blockers") if record.get("blockers") is not None else len(level_blockers),
        "moves": record.get("moves") if record.get("moves") is not None else level.get("moves", ""),
        "difficulty": safe_str(record.get("difficulty")) or safe_str(level.get("difficulty")),
        "validation_status": safe_str(record.get("validation_status")),
        "history_length": record.get("history_length") if record.get("history_length") is not None else len(history),
        "path_lengths": path_lengths,
        "tweak_reasons": tweak_reasons_for_record(record),
    }


def count_history_events(history):
    counter = Counter()
    for event in history or []:
        if not isinstance(event, dict):
            continue
        counter[safe_str(event.get("type")) or "unknown"] += 1
    return dict(counter)


def build_playtest_summary():
    latest_session = {}
    if PLAYTEST_LATEST_SESSION_PATH.is_file():
        latest_session = load_json(PLAYTEST_LATEST_SESSION_PATH)

    records = load_jsonl(PLAYTEST_EVENTS_PATH)
    normalized_records = [normalize_playtest_record(record) for record in records]

    solved = [row for row in normalized_records if row["solved"]]
    invalid_solved = [row for row in solved if row["validation_status"].upper() == "INVALID"]
    manual = [row for row in normalized_records if row["save_reason"].lower() == "manual"]
    validate = [row for row in normalized_records if row["save_reason"].lower() == "validate"]
    history_total = sum(int(to_number(row["history_length"], 0)) for row in normalized_records)
    avg_history = (history_total / len(normalized_records)) if normalized_records else 0

    def aggregate(rows, key):
        groups = {}
        for row in rows:
            group_key = safe_str(row.get(key)) or "unknown"
            groups.setdefault(group_key, []).append(row)
        out = []
        for group_key, items in sorted(groups.items(), key=lambda item: (-len(item[1]), item[0])):
            histories = [int(to_number(item.get("history_length"), 0)) for item in items]
            out.append({
                "value": group_key,
                "count": len(items),
                "solved": sum(1 for item in items if item["solved"]),
                "invalid_solved": sum(1 for item in items if item["solved"] and item["validation_status"].upper() == "INVALID"),
                "avg_history": (sum(histories) / len(histories)) if histories else 0,
            })
        return out

    tweak_candidates = sorted(
        [row for row in normalized_records if row["tweak_reasons"]],
        key=lambda row: (
            0 if "validation_invalid" in row["tweak_reasons"] else 1,
            0 if "unsolved" in row["tweak_reasons"] else 1,
            -int(to_number(row.get("history_length"), 0)),
            safe_str(row.get("level_name")).lower(),
        )
    )

    latest_history = latest_session.get("history", []) if isinstance(latest_session, dict) else []
    latest_level = latest_session.get("level", {}) if isinstance(latest_session, dict) else {}
    latest_pairs = latest_level.get("pairs") if isinstance(latest_level.get("pairs"), list) else []
    latest_blockers = latest_level.get("blockers") if isinstance(latest_level.get("blockers"), list) else []
    latest_summary = {
        "saved_at": safe_str(latest_session.get("saved_at")),
        "solved": bool(latest_session.get("solved")),
        "level_name": safe_str(latest_level.get("meta", {}).get("source_name")) or safe_str(latest_level.get("name")) or safe_str(latest_level.get("level_name")) or safe_str(latest_session.get("level_name")),
        "board": board_label(latest_level),
        "board_width": latest_level.get("board_width", ""),
        "board_height": latest_level.get("board_height", ""),
        "pairs": len(latest_pairs),
        "blockers": len(latest_blockers),
        "moves": latest_level.get("moves", ""),
        "difficulty": safe_str(latest_level.get("difficulty")),
        "history_length": len(latest_history) if isinstance(latest_history, list) else 0,
        "history_counts": count_history_events(latest_history),
        "path_lengths": {key: len(value) for key, value in (latest_session.get("paths") or {}).items()} if isinstance(latest_session, dict) else {},
    }

    latest_record = normalized_records[-1] if normalized_records else {}
    latest_summary["validation_status"] = safe_str(latest_record.get("validation_status"))
    return {
        "generated_at": "",
        "source_paths": {
            "latest_session": str(PLAYTEST_LATEST_SESSION_PATH),
            "events": str(PLAYTEST_EVENTS_PATH),
        },
        "counts": {
            "total": len(normalized_records),
            "solved": len(solved),
            "unsolved": len(normalized_records) - len(solved),
            "invalid_solved": len(invalid_solved),
            "manual": len(manual),
            "validate": len(validate),
            "avg_history": avg_history,
        },
        "latest_session": latest_summary,
        "latest_record": latest_record,
        "by_origin": aggregate(normalized_records, "origin"),
        "by_save_reason": aggregate(normalized_records, "save_reason"),
        "by_validation_status": aggregate(normalized_records, "validation_status"),
        "tweak_candidates": tweak_candidates[:15],
        "recent_sessions": normalized_records[-10:],
    }


def build_manager_rows(snapshot: dict):
    rows = []
    items_by_id = item_index(snapshot)
    saved_at = safe_str(snapshot.get("saved_at"))
    for progression_key in snapshot.get("progression_order", []):
        progression = (snapshot.get("progressions") or {}).get(progression_key) or {}
        progression_label = safe_str(progression.get("label", progression_key))
        for slot_entry in progression.get("slots", []):
            item = find_slot_item(slot_entry, items_by_id)
            level_file = safe_str(slot_entry.get("file")).strip()
            rows.append({
                "row_type": "slot",
                "saved_at": saved_at,
                "item_id": slot_entry.get("item_id", ""),
                "progression_key": progression_key,
                "progression_label": progression_label,
                "slot": slot_entry.get("slot", ""),
                "level_file": level_file,
                "level_name": display_level_name(level_file) if level_file else "",
                "difficulty": safe_str(slot_entry.get("difficulty") or item.get("difficulty")),
                "board": safe_str(slot_entry.get("board") or item.get("board")),
                "pairs": slot_entry.get("pairs", ""),
                "blockers": slot_entry.get("blockers", ""),
                "moves": slot_entry.get("moves", ""),
                "status": safe_str(item.get("status") or slot_entry.get("status")),
                "changed": "Yes" if item.get("changed") or slot_entry.get("changed") else "No",
                "notes": safe_str(item.get("notes") or slot_entry.get("notes")),
                "source_path": safe_str(item.get("source_path")),
                "saved_path": safe_str(item.get("saved_path")),
            })
    return rows


def build_rename_rows_from_progression_csvs(csv_paths: list[Path]):
    rows = []
    screenshot_lookup = load_screenshot_lookup()
    for group in build_progression_groups_from_csvs(csv_paths):
        progression_key = safe_str(group.get("progression_key")).strip()
        progression_label = compact_progression_label(group.get("progression_label") or progression_key)
        for row in group["rows"]:
            slot_value = safe_str(row.get("slot")).strip()
            if slot_value in {"", "0"}:
                continue
            level_file = safe_str(row.get("file")).strip()
            if not level_file:
                continue
            screenshot_name, screenshot_path = infer_screenshot_for_level(level_file, screenshot_lookup)
            current_name = display_level_name(level_file, screenshot_name)
            rows.append({
                "Occurrence ID": f"{progression_key}:{ordinal_label(slot_value)}",
                "Progression": progression_label,
                "Position": slot_value,
                "Ordinal": ordinal_label(slot_value),
                "Screenshot": "",
                "Current Name": current_name,
                "Target Name": "",
                "Preview Label": prefixed_level_label(slot_value, current_name),
                "Code File": level_file,
                "Level Path": safe_str(row.get("saved_path")) or f"{ROOT.name}/levels/{level_file}",
                "Screenshot File": screenshot_name,
                "Screenshot Path": screenshot_path,
                "Planned File": level_file,
                "Rename Pending": "FALSE",
                "Apply Status": "",
                "Notes": "",
                "_screenshot_path": screenshot_path,
            })
    return rows


def build_rename_rows_from_manager_rows(manager_rows: list[dict], screenshot_rows: list[dict]):
    screenshot_by_file = {
        safe_str(row.get("Code File")).strip(): row
        for row in screenshot_rows
        if safe_str(row.get("Code File")).strip()
    }
    rows = []
    for row in manager_rows:
        slot_value = safe_str(row.get("slot")).strip()
        if slot_value in {"", "0"}:
            continue
        level_file = safe_str(row.get("level_file")).strip()
        if not level_file:
            continue
        screenshot_row = screenshot_by_file.get(level_file) or {}
        screenshot_name = safe_str(screenshot_row.get("Screenshot")).strip()
        screenshot_path = safe_str(screenshot_row.get("Local Screenshot Path")).strip()
        current_name = display_level_name(level_file, screenshot_name)
        progression_key = safe_str(row.get("progression_key")).strip()
        progression_label = compact_progression_label(row.get("progression_label") or progression_key)
        rows.append({
            "Occurrence ID": f"{progression_key}:{ordinal_label(slot_value)}",
            "Progression": progression_label,
            "Position": slot_value,
            "Ordinal": ordinal_label(slot_value),
            "Screenshot": "",
            "Current Name": current_name,
            "Target Name": "",
            "Preview Label": prefixed_level_label(slot_value, current_name),
            "Code File": level_file,
            "Level Path": safe_str(row.get("saved_path")) or f"{ROOT.name}/levels/{level_file}",
            "Screenshot File": screenshot_name,
            "Screenshot Path": screenshot_path,
            "Planned File": level_file,
            "Rename Pending": "FALSE",
            "Apply Status": "",
            "Notes": "",
            "_screenshot_path": screenshot_path,
        })
    return rows


def boolish_to_yes_no(value) -> str:
    text = safe_str(value).strip().lower()
    return "Yes" if text in {"1", "true", "yes", "y"} else "No"


def find_latest_matching_csv(pattern: str) -> Path | None:
    matches = [candidate for candidate in ROOT.glob(pattern) if candidate.is_file()]
    if not matches:
        return None
    return max(matches, key=lambda candidate: candidate.stat().st_mtime)


def find_latest_matching_zip(pattern: str) -> Path | None:
    matches = [candidate for candidate in ROOT.glob(pattern) if candidate.is_file()]
    if not matches:
        return None
    return max(matches, key=lambda candidate: candidate.stat().st_mtime)


def default_progression_csvs() -> list[Path]:
    per_progression_candidates = [
        [
            ROOT / "bundles" / "original_progression_a" / "original_progression_a_progression.csv",
            ROOT / "bundles" / "original_progression_a.zip",
            ROOT / "bundles" / "progression_a" / "progression_a_progression.csv",
            find_latest_matching_zip("bundles/progression_a*.zip"),
        ],
        [
            ROOT / "bundles" / "original_progression_b" / "original_progression_b_progression.csv",
            ROOT / "bundles" / "original_progression_b.zip",
            find_latest_matching_csv("bundles/progression_b*/progression_b*_progression.csv"),
            find_latest_matching_zip("bundles/progression_b*.zip"),
        ],
        [
            ROOT / "bundles" / "original_progression_c" / "original_progression_c_progression.csv",
            ROOT / "bundles" / "original_progression_c.zip",
            find_latest_matching_csv("bundles/progression_c*/progression_c*_progression.csv"),
            find_latest_matching_zip("bundles/progression_c*.zip"),
        ],
    ]
    for csv_candidate in sorted((ROOT / "bundles").glob("progression_[a-z]/progression_[a-z]_progression.csv")):
        if csv_candidate.exists():
            per_progression_candidates.append([csv_candidate])
    # Auto-discover Procedural progression CSVs (progression_4 through progression_13+)
    for n in range(4, 20):
        key = f"progression_{n}"
        csv_candidate = ROOT / "bundles" / key / f"{key}_progression.csv"
        if csv_candidate.exists():
            per_progression_candidates.append([csv_candidate])
    selected = []
    for candidate_group in per_progression_candidates:
        match = next((candidate for candidate in candidate_group if candidate and candidate.exists()), None)
        if match:
            selected.append(match)
    return selected


def read_progression_csv_rows(csv_path: Path) -> list[dict]:
    if csv_path.suffix.lower() == ".zip":
        with zipfile.ZipFile(csv_path) as archive:
            member_name = next((name for name in archive.namelist() if name.endswith("_progression.csv")), "")
            if not member_name:
                return []
            with archive.open(member_name) as handle:
                text_stream = io.TextIOWrapper(handle, encoding="utf8", newline="")
                return list(csv.DictReader(text_stream))
    with csv_path.open(newline="", encoding="utf8") as handle:
        return list(csv.DictReader(handle))


def build_progression_groups_from_csvs(csv_paths: list[Path]) -> list[dict]:
    groups = []
    for csv_path in csv_paths:
        rows = read_progression_csv_rows(csv_path)
        if not rows:
          continue
        ordered_rows = sorted(rows, key=lambda row: int(safe_str(row.get("slot") or "0") or "0"))
        groups.append({
            "csv_path": csv_path,
            "progression_key": safe_str(ordered_rows[0].get("progression_key")),
            "progression_label": compact_progression_label(ordered_rows[0].get("progression_label") or ordered_rows[0].get("progression_key")),
            "rows": ordered_rows,
        })
    return groups


def planner_curve_formula(row_number: int) -> str:
    lookup_parts = []
    for header in PLANNER_SLOT_HEADERS:
        column_index = PLANNER_HEADERS.index(header) + 1
        lookup_parts.append(
            f"IFERROR(VLOOKUP({get_column_letter(column_index)}{row_number},'{CATALOG_SHEET_NAME}'!A:E,5,FALSE),0)"
        )
    joined = ",".join(lookup_parts)
    return f'=SPARKLINE({{{joined}}},{{"charttype","line";"linewidth",2;"empty","zero"}})'


def planner_main_row_number(index: int) -> int:
    return 2 + (index * PLANNER_ROW_BLOCK_SIZE)


def rotate_slots(values: list[str], shift: int) -> list[str]:
    if not values:
        return []
    offset = shift % len(values)
    return values[offset:] + values[:offset]


def build_catalog_rows_from_progression_csvs(csv_paths: list[Path]):
    rows = []
    seen = set()
    screenshot_lookup = load_screenshot_lookup()
    for group in build_progression_groups_from_csvs(csv_paths):
        for row in group["rows"]:
            level_file = safe_str(row.get("file")).strip()
            if not level_file or level_file in seen:
                continue
            seen.add(level_file)
            screenshot_name, _ = infer_screenshot_for_level(level_file, screenshot_lookup)
            rows.append({
                "File": level_file,
                "Display Name": display_level_name(level_file, screenshot_name),
                "Source Progression": group["progression_label"],
                "Difficulty": title_case_difficulty(row.get("difficulty")),
                "Difficulty Score": difficulty_score(row.get("difficulty"), row.get("moves"), row.get("pairs"), row.get("blockers")),
                "Board": safe_str(row.get("board")),
                "Pairs": safe_str(row.get("pairs")),
                "Blockers": safe_str(row.get("blockers")),
                "Moves": safe_str(row.get("moves")),
                "Status": safe_str(row.get("status")),
                "Screenshot": screenshot_name,
                "Saved Path": safe_str(row.get("saved_path")),
            })
    return rows


def progression_proposal_rows(group: dict) -> list[dict]:
    ordered = group["rows"]
    tutorial_row = next((row for row in ordered if safe_str(row.get("file")).strip() == "tutorial_level.json" or safe_str(row.get("slot")) == "0"), None)
    playable = [row for row in ordered if row is not tutorial_row]
    tutorial_file = safe_str(tutorial_row.get("file")) if tutorial_row else "tutorial_level.json"
    current_slots = [safe_str(row.get("file")) for row in playable][:len(PLANNER_SLOT_HEADERS)]
    family_code = progression_family_code(group["progression_key"] or group["progression_label"]) or group["progression_label"].split()[-1].upper()
    original_label = f"Original Progression {family_code}"
    def make_row(suffix: str, suggestion: str, slots: list[str]) -> dict:
        row = {
            "Proposal ID": f"original_{family_code.lower()}_{suffix}",
            "Folder Name": f"original_progression_{family_code.lower()}_{suffix}",
            "Approved": "FALSE",
            "Materialized": "FALSE",
            "Source Progression": original_label,
            "Suggestion": suggestion,
            "Tutorial File": tutorial_file,
            "Curve": "",
            "Notes": "",
            "Output Folder": "",
        }
        for index, header in enumerate(PLANNER_SLOT_HEADERS):
            row[header] = slots[index] if index < len(slots) else ""
        return row

    return [
        make_row("current", "Original Order", current_slots),
        make_row("rotate_1", "Rotate +1", rotate_slots(current_slots, 1)),
    ]


def cross_progression_proposal_rows(groups: list[dict]) -> list[dict]:
    if len(groups) < 2:
        return []
    playable_by_group = []
    tutorial_file = "tutorial_level.json"
    for group in groups:
        tutorial_row = next((row for row in group["rows"] if safe_str(row.get("file")).strip() == "tutorial_level.json" or safe_str(row.get("slot")) == "0"), None)
        if tutorial_row:
            tutorial_file = safe_str(tutorial_row.get("file")) or tutorial_file
        playable_by_group.append([
            safe_str(row.get("file"))
            for row in group["rows"]
            if row is not tutorial_row
        ][:len(PLANNER_SLOT_HEADERS)])

    def cycle_mix(start_index: int) -> list[str]:
        slots = []
        for slot_index in range(len(PLANNER_SLOT_HEADERS)):
            source_index = (slot_index + start_index) % len(playable_by_group)
            source_slots = playable_by_group[source_index]
            slots.append(source_slots[slot_index] if slot_index < len(source_slots) else "")
        return slots

    def make_row(proposal_id: str, folder_name: str, suggestion: str, slots: list[str]) -> dict:
        row = {
            "Proposal ID": proposal_id,
            "Folder Name": folder_name,
            "Approved": "FALSE",
            "Materialized": "FALSE",
            "Source Progression": "Live Ops Mixes",
            "Suggestion": suggestion,
            "Tutorial File": tutorial_file,
            "Curve": "",
            "Notes": "",
            "Output Folder": "",
        }
        for index, header in enumerate(PLANNER_SLOT_HEADERS):
            row[header] = slots[index] if index < len(slots) else ""
        return row

    return [
        make_row("liveops_mix_abc_1", "liveops_mix_abc_1", "ABC Live Ops Mix 1", cycle_mix(0)),
        make_row("liveops_mix_abc_2", "liveops_mix_abc_2", "ABC Live Ops Mix 2", cycle_mix(1)),
        make_row("liveops_mix_abc_3", "liveops_mix_abc_3", "ABC Live Ops Mix 3", cycle_mix(2)),
    ]


def build_planner_rows_from_progression_csvs(csv_paths: list[Path]):
    groups = build_progression_groups_from_csvs(csv_paths)
    rows = []
    for group in groups:
        rows.extend(progression_proposal_rows(group))
    rows.extend(cross_progression_proposal_rows(groups))
    for index, row in enumerate(rows):
        row_number = planner_main_row_number(index)
        row["Curve"] = planner_curve_formula(row_number)
    return rows


def expand_planner_rows_for_workbook(planner_rows: list[dict], catalog_rows: list[dict] | None = None):
    screenshot_lookup = load_screenshot_lookup()
    difficulty_lookup = {
        safe_str(row.get("File")).strip(): title_case_difficulty(row.get("Difficulty"))
        for row in (catalog_rows or [])
        if safe_str(row.get("File")).strip()
    }
    display_rows = []
    image_headers = ["Tutorial File", *PLANNER_SLOT_HEADERS]
    for index, row in enumerate(planner_rows):
        main_row = dict(row)
        main_row["Curve"] = planner_curve_formula(planner_main_row_number(index))
        display_rows.append(main_row)
        screenshot_row = {header: "" for header in PLANNER_HEADERS}
        screenshot_row["Suggestion"] = "Screens"
        screenshot_row["_planner_image_paths"] = {}
        for header in image_headers:
            level_file = safe_str(row.get(header)).strip()
            image_name, image_path = infer_screenshot_for_level(level_file, screenshot_lookup)
            screenshot_row[header] = image_name
            screenshot_row["_planner_image_paths"][header] = image_path
        display_rows.append(screenshot_row)
        difficulty_row = {header: "" for header in PLANNER_HEADERS}
        difficulty_row["Suggestion"] = "Difficulty"
        for header in image_headers:
            level_file = safe_str(row.get(header)).strip()
            difficulty_row[header] = difficulty_lookup.get(level_file, "")
        display_rows.append(difficulty_row)
    return display_rows


def build_rows_from_progression_csvs(csv_paths: list[Path]):
    primary_rows = []
    screenshot_rows = []
    manager_rows = []
    sequence = 1
    screenshot_lookup = load_screenshot_lookup()
    for csv_path in csv_paths:
        for row in read_progression_csv_rows(csv_path):
            level_file = safe_str(row.get("file")).strip()
            if not level_file:
                continue
            screenshot_name, screenshot_path = infer_screenshot_for_level(level_file, screenshot_lookup)
            name = display_level_name(level_file, screenshot_name)
            progression_label = compact_progression_label(row.get("progression_label") or row.get("progression_key"))
            difficulty = title_case_difficulty(row.get("difficulty"))
            changed = boolish_to_yes_no(row.get("changed"))
            note = safe_str(row.get("notes"))
            primary_rows.append({
                "Position in Progression": safe_str(row.get("slot")),
                "Level ID": f"{sequence:03d}",
                "Name/Title": prefixed_level_label(row.get("slot"), name),
                "Progression": progression_label,
                "Difficulty": difficulty,
                "Screenshot": screenshot_name,
                "Code File": level_file,
                "Need tweak": "yes" if changed == "Yes" else "no",
                "Need fixes": "FALSE",
                "Feedback": "",
                "Feedback Owner": "",
                "Playtest difficulty": "",
                "Notes": note,
                "_screenshot_path": screenshot_path,
            })
            screenshot_rows.append({
                "Level Name": name,
                "Screenshot": screenshot_name,
                "Code File": level_file,
                "Local Screenshot Path": screenshot_path,
                "Status": "FOUND" if screenshot_path and Path(screenshot_path).exists() else "MISSING",
                "_screenshot_path": screenshot_path,
            })
            manager_rows.append({
                "row_type": "slot",
                "saved_at": "",
                "item_id": "",
                "progression_key": safe_str(row.get("progression_key")),
                "progression_label": safe_str(row.get("progression_label")),
                "slot": safe_str(row.get("slot")),
                "level_file": level_file,
                "level_name": name,
                "difficulty": safe_str(row.get("difficulty")),
                "board": safe_str(row.get("board")),
                "pairs": safe_str(row.get("pairs")),
                "blockers": safe_str(row.get("blockers")),
                "moves": safe_str(row.get("moves")),
                "status": safe_str(row.get("status")),
                "changed": changed,
                "notes": note,
                "source_path": safe_str(row.get("path")),
                "saved_path": safe_str(row.get("saved_path")),
            })
            sequence += 1
    return primary_rows, screenshot_rows, manager_rows


def build_readme_rows():
    return [
        ["READ", README_SHEET_NAME, "Start here. This tab explains the workbook structure, the tab colors, and the role of each working surface.", "Do not edit manually.", ""],
        ["READ", "Blue tabs", "Reference and orientation tabs that explain the system or link out to other surfaces.", "Read only.", ""],
        ["INTERACTIVE", "Green tabs", "Editorial tabs where humans can review, plan, or stage changes.", "Edit only the intended working cells.", ""],
        ["DATA", "Orange tabs", "Generated tabs fed by toolkit exports, bundle state, or procedural snapshots.", "Do not edit manually. Regenerate from the toolkit.", ""],
        ["", "", "", "", ""],
        ["INTERACTIVE", PRIMARY_SHEET_NAME, "Main editorial review view for level order, screenshots, feedback, and playtest notes.", "Edit only the review columns such as Need fixes, Feedback, Feedback Owner, and Playtest difficulty.", ""],
        ["INTERACTIVE", PLANNER_SHEET_NAME, "Plan and approve Live Ops Mixes before materialization.", "Edit proposal, approval, slot, tutorial, and notes fields only.", ""],
        ["INTERACTIVE", RENAME_SHEET_NAME, "Stage level renames before applying them locally.", "Edit Target Name and Notes, then run the rename action.", ""],
        ["DATA", CATALOG_SHEET_NAME, "Canonical level catalog generated from bundle state.", "Do not edit manually.", ""],
        ["DATA", "Level Manager state", "Machine-generated sync surface used by toolkit and export flows.", "Do not edit manually.", ""],
        ["DATA", "Procedural learning", "Snapshot of approved and rejected procedural feedback for review and reporting.", "Do not edit manually.", ""],
        ["", "", "", "", ""],
        ["ACTION", "Open Web Toolkit", "Open the local toolkit UI.", "Runs on the local server.", f"{TOOLKIT_BASE_URL}/"],
        ["ACTION", "Sync Spreadsheet", "Rebuild the canonical workbook and payload, then push the live Google Sheet.", "Runs on the local server.", f"{TOOLKIT_BASE_URL}/api/action/sync-spreadsheet"],
        ["ACTION", "Regenerate Payload", "Rebuild the canonical workbook and payload locally without a live push.", "Runs on the local server.", f"{TOOLKIT_BASE_URL}/api/action/generate-payload"],
        ["ACTION", "Apply Level Renames", "Apply staged rename rows back to the canonical repo files.", "Runs on the local server.", f"{TOOLKIT_BASE_URL}/api/action/apply-level-renames"],
        ["ACTION", "Materialize Mixes", "Export approved Live Ops Mixes into bundle output.", "Runs on the local server.", f"{TOOLKIT_BASE_URL}/api/action/materialize-mixes"],
        ["ACTION", "Validate All Levels", "Run the current all-level validation action.", "Runs on the local server.", f"{TOOLKIT_BASE_URL}/api/action/validate-levels"],
        ["ACTION", "Backup Progressions", "Create a local progression backup before major changes.", "Runs on the local server.", f"{TOOLKIT_BASE_URL}/api/action/backup-progressions"],
        ["", "", "", "", ""],
        ["READ", "Game URL", "Playable build reference.", "Reference only.", "https://king-ccss-minigame-hungrybears.firebaseapp.com/"],
        ["READ", "Team board", "Shared board for wider team context.", "Reference only.", "https://miro.com/app/board/uXjVJTMkKXk=/"],
    ]


def reset_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def ensure_only_canonical_sheets(wb: Workbook):
    for title in list(wb.sheetnames):
        del wb[title]


def write_rows(ws, headers, rows):
    ws.delete_rows(1, ws.max_row or 1)
    ws.append(headers)
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = value
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append([row.get(key, "") for key in headers])
    ws.freeze_panes = "A2"


def set_widths(ws, widths: dict[int, float]):
    for index, value in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = value


def fit_text_columns(ws, headers, max_width=44):
    for index, _header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        width = ws.column_dimensions[letter].width or 0
        if width:
            continue
        longest = max(len(safe_str(ws.cell(row=row, column=index).value)) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), max_width)


def embed_sheet_images(ws, rows, image_column_index: int, path_key: str, size_px: int = 120):
    ws.column_dimensions[get_column_letter(image_column_index)].width = 20
    for row_index, row in enumerate(rows, start=2):
        image_path = resolve_project_path(row.get(path_key))
        if not str(image_path) or not image_path.is_file():
            continue
        ws.cell(row=row_index, column=image_column_index).value = ""
        image = XLImage(str(image_path))
        image.width = size_px
        image.height = size_px
        image.anchor = f"{get_column_letter(image_column_index)}{row_index}"
        ws.add_image(image)
        ws.row_dimensions[row_index].height = 95


def embed_planner_images(ws, rows, headers, size_px: int = 72):
    image_headers = ["Tutorial File", *PLANNER_SLOT_HEADERS]
    header_lookup = {header: index for index, header in enumerate(headers, start=1)}
    for row_index, row in enumerate(rows, start=2):
        image_paths = row.get("_planner_image_paths") or {}
        if not image_paths:
            continue
        ws.row_dimensions[row_index].height = 62
        for header in image_headers:
            image_path = resolve_project_path(image_paths.get(header))
            if not str(image_path) or not image_path.is_file():
                continue
            column_index = header_lookup.get(header)
            if not column_index:
                continue
            ws.cell(row=row_index, column=column_index).value = ""
            image = XLImage(str(image_path))
            image.width = size_px
            image.height = size_px
            image.anchor = f"{get_column_letter(column_index)}{row_index}"
            ws.add_image(image)


def format_planner_workbook(ws, headers, rows):
    image_headers = ["Tutorial File", *PLANNER_SLOT_HEADERS]
    header_lookup = {header: index for index, header in enumerate(headers, start=1)}
    for block_start in range(2, len(rows) + 2, PLANNER_ROW_BLOCK_SIZE):
        difficulty_row_index = block_start + 2
        if difficulty_row_index > ws.max_row:
            continue
        ws.row_dimensions[difficulty_row_index].height = 24
        for header in image_headers:
            column_index = header_lookup.get(header)
            if not column_index:
                continue
            cell = ws.cell(row=difficulty_row_index, column=column_index)
            key = safe_str(cell.value).strip().upper()
            cell.fill = DIFFICULTY_FILLS.get(key, DIFFICULTY_DEFAULT_FILL)
            cell.font = Font(bold=True, color=DIFFICULTY_FONT_COLORS.get(key, DIFFICULTY_DEFAULT_FONT_COLOR))
            cell.alignment = Alignment(horizontal="center", vertical="center")


def format_difficulty_columns_workbook(ws, headers, difficulty_headers=None):
    target_headers = difficulty_headers or ["Difficulty", "Playtest difficulty"]
    for header in target_headers:
        if header not in headers:
            continue
        column_index = headers.index(header) + 1
        for row_index in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=column_index)
            key = safe_str(cell.value).strip().upper()
            if key not in DIFFICULTY_FILLS:
                continue
            cell.fill = DIFFICULTY_FILLS[key]
            cell.font = Font(bold=True, color=DIFFICULTY_FONT_COLORS[key])
            cell.alignment = Alignment(horizontal="center", vertical="center")


def workbook_tab_group(sheet_name: str) -> str:
    if sheet_name == README_SHEET_NAME:
        return "read"
    if sheet_name in {PRIMARY_SHEET_NAME, PLANNER_SHEET_NAME, RENAME_SHEET_NAME}:
        return "interactive"
    return "data"


def set_workbook_tab_color(ws, group: str):
    color = WORKBOOK_TAB_COLORS.get(group)
    if color:
        ws.sheet_properties.tabColor = color


def build_payload(primary_rows, screenshot_rows, manager_rows, readme_headers, readme_rows, catalog_rows=None, planner_rows=None, rename_rows=None, playtest_summary=None):
    screenshot_matrix = [[row.get(key, "") for key in SCREENSHOT_HEADERS] for row in screenshot_rows]
    return {
        "readmeSheetName": README_SHEET_NAME,
        "readmeHeaders": readme_headers,
        "readmeRows": readme_rows,
        "primarySheetName": PRIMARY_SHEET_NAME,
        "primaryHeaders": PRIMARY_HEADERS,
        "primaryRows": [[row.get(key, "") for key in PRIMARY_HEADERS] for row in primary_rows],
        # Keep both spellings until every consumer is moved to the canonical key.
        "screenshotHeaders": SCREENSHOT_HEADERS,
        "screenshotRows": screenshot_matrix,
        "screenshootRows": screenshot_matrix,
        "dbHeaders": DB_HEADERS,
        "dbRows": [[row.get(key, "") for key in DB_HEADERS] for row in manager_rows],
        "catalogSheetName": CATALOG_SHEET_NAME,
        "catalogHeaders": CATALOG_HEADERS,
        "catalogRows": [[row.get(key, "") for key in CATALOG_HEADERS] for row in (catalog_rows or [])],
        "plannerSheetName": PLANNER_SHEET_NAME,
        "plannerHeaders": PLANNER_HEADERS,
        "plannerRows": [[row.get(key, "") for key in PLANNER_HEADERS] for row in (planner_rows or [])],
        "renameSheetName": RENAME_SHEET_NAME,
        "renameHeaders": RENAME_HEADERS,
        "renameRows": [[row.get(key, "") for key in RENAME_HEADERS] for row in (rename_rows or [])],
        "linksHeaders": readme_headers,
        "linksRows": readme_rows,
        "playtestSummary": playtest_summary or {},
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE))
    parser.add_argument("--snapshot")
    parser.add_argument("--progression-csv", action="append", default=[])
    parser.add_argument("--use-default-progressions", action="store_true")
    parser.add_argument("--from-bundles", action="store_true")
    parser.add_argument("--payload-output")
    args = parser.parse_args()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    readme_headers = README_HEADERS
    readme_rows = build_readme_rows()
    snapshot = None
    csv_paths = []
    if args.use_default_progressions or args.from_bundles:
        csv_paths = default_progression_csvs()
        if not csv_paths:
            parser.error("No default progression CSVs were found under bundles/.")
    elif args.progression_csv:
        csv_paths = [Path(candidate).resolve() for candidate in args.progression_csv]
    elif args.snapshot:
        snapshot = load_json(Path(args.snapshot))
    else:
        parser.error("Provide --snapshot, --progression-csv, or --use-default-progressions.")

    catalog_rows = []
    planner_rows = []
    rename_rows = []
    playtest_summary = build_playtest_summary()
    if csv_paths:
        primary_rows, screenshot_rows, manager_rows = build_rows_from_progression_csvs(csv_paths)
        catalog_rows = build_catalog_rows_from_progression_csvs(csv_paths)
        planner_rows = build_planner_rows_from_progression_csvs(csv_paths)
        rename_rows = build_rename_rows_from_progression_csvs(csv_paths)
    else:
        primary_rows, screenshot_rows = build_primary_rows(snapshot)
        manager_rows = build_manager_rows(snapshot)
        rename_rows = build_rename_rows_from_manager_rows(manager_rows, screenshot_rows)
    if playtest_summary:
        playtest_summary["generated_at"] = datetime.now(timezone.utc).isoformat()
    payload = build_payload(primary_rows, screenshot_rows, manager_rows, readme_headers, readme_rows, catalog_rows, planner_rows, rename_rows, playtest_summary)

    wb = reset_workbook()
    ensure_only_canonical_sheets(wb)

    readme_ws = wb.create_sheet(README_SHEET_NAME)
    write_rows(readme_ws, readme_headers, [{readme_headers[index]: row[index] for index in range(len(readme_headers))} for row in readme_rows])
    set_widths(readme_ws, {1: 14, 2: 24, 3: 48, 4: 38, 5: 30})
    set_workbook_tab_color(readme_ws, workbook_tab_group(README_SHEET_NAME))

    after_ws = wb.create_sheet(PRIMARY_SHEET_NAME)
    write_rows(after_ws, PRIMARY_HEADERS, primary_rows)
    set_widths(after_ws, {1: 18, 2: 12, 3: 22, 4: 18, 5: 12, 6: 20, 7: 24, 8: 12, 9: 12, 10: 32, 11: 18, 12: 18, 13: 24})
    fit_text_columns(after_ws, PRIMARY_HEADERS)
    embed_sheet_images(after_ws, primary_rows, 6, "_screenshot_path")
    format_difficulty_columns_workbook(after_ws, PRIMARY_HEADERS)
    set_workbook_tab_color(after_ws, workbook_tab_group(PRIMARY_SHEET_NAME))

    if planner_rows:
        planner_ws = wb.create_sheet(PLANNER_SHEET_NAME)
        planner_display_rows = expand_planner_rows_for_workbook(planner_rows, catalog_rows)
        write_rows(planner_ws, PLANNER_HEADERS, planner_display_rows)
        set_widths(planner_ws, {1: 24, 2: 24, 3: 12, 4: 14, 5: 18, 6: 16, 7: 22, 18: 18, 19: 24, 20: 28})
        fit_text_columns(planner_ws, PLANNER_HEADERS, max_width=24)
        format_planner_workbook(planner_ws, PLANNER_HEADERS, planner_display_rows)
        set_workbook_tab_color(planner_ws, workbook_tab_group(PLANNER_SHEET_NAME))

    if rename_rows:
        rename_ws = wb.create_sheet(RENAME_SHEET_NAME)
        write_rows(rename_ws, RENAME_HEADERS, rename_rows)
        set_widths(rename_ws, {1: 20, 2: 18, 3: 10, 4: 10, 5: 18, 6: 20, 7: 20, 8: 22, 9: 22, 10: 28, 11: 18, 12: 28, 13: 22, 14: 14, 15: 18, 16: 18})
        fit_text_columns(rename_ws, RENAME_HEADERS, max_width=30)
        set_workbook_tab_color(rename_ws, workbook_tab_group(RENAME_SHEET_NAME))

    if catalog_rows:
        catalog_ws = wb.create_sheet(CATALOG_SHEET_NAME)
        write_rows(catalog_ws, CATALOG_HEADERS, catalog_rows)
        fit_text_columns(catalog_ws, CATALOG_HEADERS, max_width=28)
        format_difficulty_columns_workbook(catalog_ws, CATALOG_HEADERS, ["Difficulty"])
        set_workbook_tab_color(catalog_ws, workbook_tab_group(CATALOG_SHEET_NAME))

    manager_ws = wb.create_sheet("Level Manager state")
    write_rows(manager_ws, DB_HEADERS, manager_rows)
    fit_text_columns(manager_ws, DB_HEADERS, max_width=36)
    set_workbook_tab_color(manager_ws, workbook_tab_group("Level Manager state"))

    wb.save(output_path)
    if args.payload_output:
        Path(args.payload_output).write_text(json.dumps(payload, indent=2))
    if playtest_summary:
        PLAYTEST_SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
        PLAYTEST_SUMMARY_PATH.write_text(json.dumps(playtest_summary, indent=2), encoding="utf8")
    print(output_path)


if __name__ == "__main__":
    main()

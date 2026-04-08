#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill


VALID_EXTENSIONS = {".png", ".jpg", ".jpeg"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create an Excel workbook with embedded screenshots."
    )
    parser.add_argument(
        "--source-dir",
        default="levels/screenshots",
        help="Directory containing image files.",
    )
    parser.add_argument(
        "--output",
        default="output/spreadsheet/level_screenshots_embedded.xlsx",
        help="Output .xlsx path.",
    )
    parser.add_argument(
        "--max-width",
        type=int,
        default=220,
        help="Maximum embedded image width in pixels.",
    )
    parser.add_argument(
        "--max-height",
        type=int,
        default=160,
        help="Maximum embedded image height in pixels.",
    )
    return parser.parse_args()


def fit_dimensions(width: int, height: int, max_width: int, max_height: int) -> tuple[int, int]:
    scale = min(max_width / width, max_height / height, 1)
    return max(1, int(width * scale)), max(1, int(height * scale))


def main() -> None:
    args = parse_args()
    source_dir = Path(args.source_dir)
    output_path = Path(args.output)

    if not source_dir.exists():
      raise SystemExit(f"Source directory not found: {source_dir}")

    image_paths = sorted(
        path for path in source_dir.iterdir() if path.is_file() and path.suffix.lower() in VALID_EXTENSIONS
    )
    if not image_paths:
        raise SystemExit(f"No images found in: {source_dir}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Screenshots"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Name", "Preview", "Filename", "Relative Path"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 52
    ws.row_dimensions[1].height = 24

    for row_index, image_path in enumerate(image_paths, start=2):
        stem = image_path.stem
        ws.cell(row=row_index, column=1, value=stem)
        ws.cell(row=row_index, column=3, value=image_path.name)
        ws.cell(row=row_index, column=4, value=str(image_path))

        preview = XLImage(str(image_path))
        preview.width, preview.height = fit_dimensions(
            preview.width, preview.height, args.max_width, args.max_height
        )
        ws.add_image(preview, f"B{row_index}")

        # Approximate row height so the embedded image is visible without manual resize.
        ws.row_dimensions[row_index].height = max(120, int(preview.height * 0.78))

        for column in ("A", "C", "D"):
            ws[f"{column}{row_index}"].alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)
    print(f"Created workbook: {output_path}")
    print(f"Embedded images: {len(image_paths)}")


if __name__ == "__main__":
    main()
